import time
import requests
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font
import os
import logging

# === LOG SETUP ===
LOG_FILE = 'update_log.txt'
logging.basicConfig(filename=LOG_FILE, level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

# === CONFIG ===
API_TOKEN = 'eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiIsImtpZCI6IjI4YTMxOGY3LTAwMDAtYTFlYi03ZmExLTJjNzQzM2M2Y2NhNSJ9.eyJpc3MiOiJzdXBlcmNlbGwiLCJhdWQiOiJzdXBlcmNlbGw6Z2FtZWFwaSIsImp0aSI6IjlhZThmMWMzLTI4OGItNDZkMS1iZjk4LTZkMjQ0MDQ2MTYyZSIsImlhdCI6MTc0ODQ1NjM1MSwic3ViIjoiZGV2ZWxvcGVyL2NjYzBiN2Q5LTgzMjQtN2Q4ZS02YWQ1LTkyNjVmYzc4M2VlZCIsInNjb3BlcyI6WyJjbGFzaCJdLCJsaW1pdHMiOlt7InRpZXIiOiJkZXZlbG9wZXIvc2lsdmVyIiwidHlwZSI6InRocm90dGxpbmcifSx7ImNpZHJzIjpbIjQ5LjM3LjIyMy4xNTciXSwidHlwZSI6ImNsaWVudCJ9XX0.YncivW25SQaq3JhrcuqC3qGJSUIG5JIc6yHubTDrhP8xxGjvE4Xyvt5NM2qPJ0ZqVovTn7gTVhYmja5zG1fdsw'
CLAN_TAG = '#PQJJQ2PG'
HEADERS = {
    "Authorization": f"Bearer {API_TOKEN}",
    "Accept": "application/json"
}
API_BASE = 'https://api.clashofclans.com/v1'
FILENAME = 'live_war_auto_update.xlsx'

# === TIME FORMATTERS ===
def format_time(time_str):
    # Converts ISO8601 format to readable datetime string
    return datetime.strptime(time_str, "%Y%m%dT%H%M%S.%fZ").strftime('%Y-%m-%d %H:%M:%S')

def get_remaining_time(end_time_str):
    end_dt = datetime.strptime(end_time_str, "%Y%m%dT%H%M%S.%fZ").replace(tzinfo=timezone.utc)
    now_utc = datetime.now(timezone.utc)
    remaining = end_dt - now_utc
    return "00:00:00" if remaining.total_seconds() < 0 else str(remaining).split('.')[0]

# === API CALL without IP forcing ===
def get_current_war(clan_tag):
    url = f"{API_BASE}/clans/{clan_tag.replace('#', '%23')}/currentwar"
    response = requests.get(url, headers=HEADERS, timeout=10)
    response.raise_for_status()
    return response.json()

# === PLAYER STATS ===
def extract_player_data(clan):
    data = []
    stars_total = 0
    destruction_total = 0
    townhall_sum = 0

    for i, member in enumerate(clan.get('members', []), 1):
        name = member.get('name', '')
        tag = member.get('tag', '')
        th = member.get('townhallLevel', 0)
        map_pos = member.get('mapPosition', '')
        townhall_sum += th
        total_stars = 0
        total_destruction = 0
        att_1 = att_2 = {}

        attacks = sorted(member.get('attacks', []), key=lambda x: x.get('order', 0))
        if len(attacks) > 0:
            att_1 = attacks[0]
            total_stars += att_1.get('stars', 0)
            total_destruction += att_1.get('destructionPercentage', 0)
        if len(attacks) > 1:
            att_2 = attacks[1]
            total_stars += att_2.get('stars', 0)
            total_destruction += att_2.get('destructionPercentage', 0)

        stars_total += total_stars
        destruction_total += total_destruction

        points = total_stars * 10 + total_destruction * 0.1
        row = [
            i, tag, name, th, map_pos, total_stars, round(total_destruction, 1),
            att_1.get('defenderTag', ''), att_1.get('stars', ''), att_1.get('destructionPercentage', ''),
            att_2.get('defenderTag', ''), att_2.get('stars', ''), att_2.get('destructionPercentage', ''),
            round(points, 1)
        ]
        data.append(row)

    total_members = len(clan.get('members', [])) or 1  # avoid div zero
    summary = {
        'total_players': total_members,
        'clan_tag': clan.get('tag', ''),
        'clan_name': clan.get('name', ''),
        'avg_th': round(townhall_sum / total_members, 1),
        'clan_stars': stars_total,
        'avg_destruction': round(destruction_total / total_members, 1),
        'tips': "Target low bases with 2nd attackers. Aim 2⭐+ high %."
    }
    return data, summary

# === EXCEL WRITE ===
def write_to_excel(our_data, opp_data, our_summary, opp_summary, war_info):
    if os.path.exists(FILENAME):
        try:
            os.remove(FILENAME)
        except PermissionError:
            raise PermissionError(f"❌ Cannot overwrite '{FILENAME}'. Please close the file and try again.")

    wb = openpyxl.Workbook()

    # Our Clan Sheet
    ws_our = wb.active
    ws_our.title = "Our Clan War Report"

    ws_our.append(["📅 War Timing Info"])
    for k, v in war_info.items():
        ws_our.append([k, v])
    ws_our.append([])

    headers = ["No", "Tag", "Name", "Townhall", "Map Position", "Total Stars", "Total Destruction",
               "1st Attack Target", "1st Attack Stars", "1st Attack Destruction",
               "2nd Attack Target", "2nd Attack Stars", "2nd Attack Destruction", "Attacker Points"]
    ws_our.append(headers)
    for cell in ws_our[ws_our.max_row]:
        cell.font = Font(bold=True)

    for row in our_data:
        ws_our.append(row)

    ws_our.append([])
    ws_our.append(["Number of players", "Clan tag", "Clan name", "Average Townhall", "Strategy Tips", "Clan Stars", "Average Destruction"])
    ws_our.append([
        our_summary['total_players'],
        our_summary['clan_tag'],
        our_summary['clan_name'],
        our_summary['avg_th'],
        our_summary['tips'],
        our_summary['clan_stars'],
        our_summary['avg_destruction']
    ])

    # Opponent Clan Sheet
    ws_opp = wb.create_sheet(title="Opponent Clan War Report")
    ws_opp.append(["Opponent Clan War Report"])
    ws_opp.append([])

    headers_opp = ["No", "Tag", "Name", "Townhall", "Map Position", "Total Stars", "Total Destruction",
                   "1st Attack Target", "1st Attack Stars", "1st Attack Destruction",
                   "2nd Attack Target", "2nd Attack Stars", "2nd Attack Destruction", "Attacker Points"]
    ws_opp.append(headers_opp)
    for cell in ws_opp[ws_opp.max_row]:
        cell.font = Font(bold=True)

    for row in opp_data:
        ws_opp.append(row)

    ws_opp.append([])
    ws_opp.append(["Number of players", "Clan tag", "Clan name", "Average Townhall", "Strategy Tips", "Clan Stars", "Average Destruction"])
    ws_opp.append([
        opp_summary['total_players'],
        opp_summary['clan_tag'],
        opp_summary['clan_name'],
        opp_summary['avg_th'],
        opp_summary['tips'],
        opp_summary['clan_stars'],
        opp_summary['avg_destruction']
    ])

    wb.save(FILENAME)
    return FILENAME

# === MAIN LOOP ===
if __name__ == "__main__":
    logging.info("=== War Auto Update Script Started ===")
    while True:
        try:
            logging.info("Fetching latest war data...")
            war = get_current_war(CLAN_TAG)
            state = war.get('state', '')
            if state not in ['inWar', 'warEnded']:
                logging.warning("No war data available yet or war not started.")
            else:
                our_clan = war['clan']
                opp_clan = war['opponent']
                # Swap if our clan is actually opponent (defensive check)
                if our_clan.get('tag', '').upper() != CLAN_TAG.upper():
                    our_clan, opp_clan = opp_clan, our_clan

                war_info = {
                    "Preparation Start": format_time(war['preparationStartTime']),
                    "War Start Time": format_time(war['startTime']),
                    "War End Time": format_time(war['endTime']),
                    "War State": state
                }
                if state == "inWar":
                    war_info["⏳ Time Until War Ends"] = get_remaining_time(war['endTime'])

                our_data, our_summary = extract_player_data(our_clan)
                opp_data, opp_summary = extract_player_data(opp_clan)

                file_path = write_to_excel(our_data, opp_data, our_summary, opp_summary, war_info)
                logging.info(f"Excel updated: {file_path}")

        except PermissionError as pe:
            logging.error(str(pe))
        except requests.exceptions.HTTPError as http_err:
            logging.error(f"HTTP error occurred: {http_err}")
        except requests.exceptions.ConnectionError as conn_err:
            logging.error(f"Connection error: {conn_err}")
        except Exception as e:
            logging.error(f"Unexpected error: {e}")

        time.sleep(60)
